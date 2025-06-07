import openpyxl
from pathlib import Path
from sqlalchemy import select, func, case, and_
from sqlalchemy.orm import selectinload, aliased
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from backend.database.database import db_session
from backend.database.models.elective import Elective
from backend.database.models.group import Group
from backend.database.models.manager import Manager
from backend.database.models.report import Report
from backend.database.models.student import Student, student_group
from backend.database.models.transfer import Transfer, TransferStatus
from backend.logic.services.log_service.orm import DatabaseLogger

log = DatabaseLogger(__name__)


class ReportService:
    """
    Сервис для генерации Excel-отчетов.
    Разделяет логику на три части:
    1. Получение данных (SQL-запросы).
    2. Создание и наполнение Excel-файла.
    3. Вспомогательные функции для работы с файлами и стилями.
    """

    REPORTS_DIR = Path("reports")

    def __init__(self):
        self.REPORTS_DIR.mkdir(exist_ok=True)
        # ✅ Определяем стили для статусов один раз, чтобы не создавать их в цикле
        self.status_styles = {
            "Подтверждено": {
                "fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                "font": Font(color="006100")
            },
            "Отклонено": {
                "fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                "font": Font(color="9C0006")
            },
            "Не обработано": {
                "fill": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
                "font": Font(color="9C6500")
            }
        }

    # region --- Основная логика генерации ---

    @db_session
    async def generate_report(self, report_type: str, db: AsyncSession) -> Report:
        """Главный метод, который создает отчет в зависимости от типа."""
        report = Report(report_type=report_type, status="generating")
        db.add(report)
        await db.flush()

        try:
            handler = self._get_report_handler(report_type)
            await handler(report, db)

            report.status = "completed"
            report.download_url = f"/reports/{report.id}"
            await db.commit()
            return report

        except Exception as e:
            log.error(f"Ошибка генерации отчета {report.id}: {e}", exc_info=True)
            report.status = "failed"
            await db.commit()
            raise

    def _get_report_handler(self, report_type: str):
        """Возвращает соответствующий обработчик для типа отчета."""
        handlers = {
            "student_transfers": self._generate_student_transfers_report,
            "elective_statistics": self._generate_elective_statistics_report,
            "manager_actions": self._generate_manager_actions_report,
        }
        handler = handlers.get(report_type)
        if not handler:
            raise ValueError(f"Неизвестный тип отчета: {report_type}")
        return handler

    async def _generate_student_transfers_report(self, report: Report, db: AsyncSession):
        """Генерирует отчет по переводам студентов."""
        headers = [
            "ID Перевода", "ФИО Студента", "УП Студента",
            "Из электива", "Из групп", "В электив", "В группы",
            "Статус", "Приоритет", "Дата создания", "Дата изменения статуса", "Менеджер"
        ]
        transfers_data = await self._fetch_student_transfers_data(db)
        wb, ws = self._create_workbook_with_headers("Переводы студентов", headers)
        self._render_student_transfers_sheet(ws, transfers_data)
        self._finalize_report(wb, ws, len(headers), report.id)

    def _render_student_transfers_sheet(self, ws: Worksheet, transfers: list[Transfer]):
        """Наполняет лист данными по переводам студентов."""
        for row_idx, transfer in enumerate(transfers, 2):
            # ✅ 1.1: Порядок соответствует новому порядку заголовков
            ws.cell(row=row_idx, column=1).value = transfer.id
            ws.cell(row=row_idx, column=2).value = transfer.student.fio
            ws.cell(row=row_idx, column=3).value = transfer.student.sp_profile
            ws.cell(row=row_idx, column=4).value = transfer.from_elective.name
            ws.cell(row=row_idx, column=5).value = ", ".join([g.name for g in transfer.groups_from])
            ws.cell(row=row_idx, column=6).value = transfer.to_elective.name
            ws.cell(row=row_idx, column=7).value = ", ".join([g.name for g in transfer.groups_to])

            status_cell = ws.cell(row=row_idx, column=8)
            status_cell.value = transfer.status_ru
            # Подкрашиваем ячейку статуса
            if style := self.status_styles.get(transfer.status_ru):
                status_cell.fill = style["fill"]
                status_cell.font = style["font"]

            ws.cell(row=row_idx, column=9).value = transfer.priority
            ws.cell(row=row_idx, column=10).value = transfer.created_at
            ws.cell(row=row_idx, column=11).value = transfer.updated_at
            ws.cell(row=row_idx, column=12).value = transfer.manager.name if transfer.manager else "N/A"

    async def _fetch_student_transfers_data(self, db: AsyncSession) -> list[Transfer]:
        """Получает данные о переводах с подгрузкой связанных сущностей."""
        FromElective = aliased(Elective)
        status_case = case(
            (Transfer.status == TransferStatus.approved, "Подтверждено"),
            (Transfer.status == TransferStatus.rejected, "Отклонено"),
            else_="Не обработано"
        ).label("status_ru")

        query = (
            select(Transfer, status_case)
            .join(Student, Transfer.student_id == Student.id)
            .join(FromElective, Transfer.from_elective_id == FromElective.id)
            .options(
                selectinload(Transfer.student), selectinload(Transfer.manager),
                selectinload(Transfer.from_elective), selectinload(Transfer.to_elective),
                selectinload(Transfer.groups_from), selectinload(Transfer.groups_to)
            ).order_by(Student.fio, FromElective.name)
        )
        result = await db.execute(query)
        transfers = []
        for transfer_obj, status_ru_val in result.all():
            transfer_obj.status_ru = status_ru_val
            transfers.append(transfer_obj)
        return transfers


    async def _generate_elective_statistics_report(self, report: Report, db: AsyncSession):
        """Генерирует статистику по элективам."""
        headers = [
            "ID", "Название", "Кластер", "Текущее кол-во студентов",
            "Студентов в заявках", "Сколько хотели ИЗ", "Сколько хотели В",
            "Сколько получилось В"
        ]
        stats_data = await self._fetch_elective_statistics_data(db)
        wb, ws = self._create_workbook_with_headers("Статистика элективов", headers)
        self._render_elective_statistics_sheet(ws, stats_data)
        self._finalize_report(wb, ws, len(headers), report.id)

    def _render_elective_statistics_sheet(self, ws: Worksheet, stats: list):
        """Наполняет лист данными по статистике элективов."""
        for row_idx, stat in enumerate(stats, 2):
            ws.cell(row=row_idx, column=1).value = stat.id
            ws.cell(row=row_idx, column=2).value = stat.name
            ws.cell(row=row_idx, column=3).value = stat.cluster
            ws.cell(row=row_idx, column=4).value = stat.current_students
            ws.cell(row=row_idx, column=5).value = stat.unique_students_in_transfers
            ws.cell(row=row_idx, column=6).value = stat.wanted_from_count
            ws.cell(row=row_idx, column=7).value = stat.wanted_to_count

            ws.cell(row=row_idx, column=8).value = stat.approved_to_count

    async def _fetch_elective_statistics_data(self, db: AsyncSession):
        """Получает статистику по элективам с обновленной логикой подсчетов и сортировкой."""
        unique_students_in_transfers_agg = func.count(func.distinct(Transfer.student_id))

        query = (
            select(
                Elective.id, Elective.name, Elective.cluster,
                func.count(func.distinct(student_group.c.student_id)).label("current_students"),
                unique_students_in_transfers_agg.label("unique_students_in_transfers"),
                func.count(func.distinct(
                    case((Transfer.from_elective_id == Elective.id, Transfer.student_id), else_=None)
                )).label("wanted_from_count"),
                func.count(func.distinct(
                    case((Transfer.to_elective_id == Elective.id, Transfer.student_id), else_=None)
                )).label("wanted_to_count"),
                func.count(func.distinct(
                    case((and_(Transfer.to_elective_id == Elective.id, Transfer.status == TransferStatus.approved),
                          Transfer.student_id), else_=None)
                )).label("approved_to_count"),
            )
            .outerjoin(Group, Elective.id == Group.elective_id)
            .outerjoin(student_group, Group.id == student_group.c.group_id)
            .outerjoin(
                Transfer,
                (Elective.id == Transfer.from_elective_id) | (Elective.id == Transfer.to_elective_id)
            )
            .group_by(Elective.id, Elective.name, Elective.cluster)
            .order_by(unique_students_in_transfers_agg.desc(), Elective.name)
        )
        result = await db.execute(query)
        return result.all()

    async def _generate_manager_actions_report(self, report: Report, db: AsyncSession):
        """Генерирует отчет по действиям менеджеров."""
        headers = ["Менеджер", "Всего заявок", "Одобрено", "Отклонено", "Последнее действие"]
        actions_data = await self._fetch_manager_actions_data(db)
        wb, ws = self._create_workbook_with_headers("Действия менеджеров", headers)
        self._render_manager_actions_sheet(ws, actions_data)
        self._finalize_report(wb, ws, len(headers), report.id)

    def _render_manager_actions_sheet(self, ws: Worksheet, actions: list):
        """Наполняет лист данными по действиям менеджеров."""
        for row_idx, action in enumerate(actions, 2):
            ws.cell(row=row_idx, column=1).value = action.manager_name
            ws.cell(row=row_idx, column=2).value = action.total_transfers
            ws.cell(row=row_idx, column=3).value = action.approved_transfers
            ws.cell(row=row_idx, column=4).value = action.rejected_transfers
            ws.cell(row=row_idx, column=5).value = action.last_action

    async def _fetch_manager_actions_data(self, db: AsyncSession):
        """Получает данные для отчета по действиям менеджеров."""
        query = (
            select(
                Manager.name.label("manager_name"),
                func.count(func.distinct(Transfer.id)).label("total_transfers"),
                func.count(func.distinct(
                    case((Transfer.status == TransferStatus.approved, Transfer.id), else_=None)
                )).label("approved_transfers"),
                func.count(func.distinct(
                    case((Transfer.status == TransferStatus.rejected, Transfer.id), else_=None)
                )).label("rejected_transfers"),
                func.max(Transfer.updated_at).label("last_action"),
            )
            .outerjoin(Transfer, Manager.id == Transfer.manager_id)
            .group_by(Manager.id, Manager.name)
            .order_by(func.count(func.distinct(Transfer.id)).desc())
        )
        result = await db.execute(query)
        return result.all()

    def _create_workbook_with_headers(self, title: str, headers: list) -> tuple[Workbook, Worksheet]:
        """Создает книгу Excel, активный лист и стилизованные заголовки."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title

        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        return wb, ws

    def _adjust_column_widths(self, ws: Worksheet, num_columns: int):
        """Автоматически настраивает ширину столбцов."""
        for i in range(1, num_columns + 1):
            ws.column_dimensions[get_column_letter(i)].auto_size = True

    def _get_report_path(self, report_id: int) -> Path:
        """Получает путь к файлу отчета."""
        return self.REPORTS_DIR / f"report_{report_id}.xlsx"

    def _finalize_report(self, wb: Workbook, ws: Worksheet, num_columns: int, report_id: int):
        """Объединяет финальные шаги: настройка ширины и сохранение файла."""
        self._adjust_column_widths(ws, num_columns)
        file_path = self._get_report_path(report_id)
        wb.save(file_path)

    @db_session
    async def get_available_reports(self, db: AsyncSession) -> list[Report]:
        """Получает список доступных отчетов"""
        result = await db.execute(select(Report).order_by(Report.created_at.desc()))
        return list(result.scalars().all())

    @db_session
    async def get_report(self, report_id: int, db: AsyncSession) -> Report:
        """Получает информацию о конкретном отчете"""
        report = await db.get(Report, report_id)
        if report and report.status == "completed":
            report.file_path = self._get_report_path(report.id)
        return report
